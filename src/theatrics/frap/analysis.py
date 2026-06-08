from __future__ import annotations

import numpy as np
from scipy.optimize import curve_fit
from scipy.special import i0, i1
from skimage.draw import disk
import xml.etree.ElementTree as ET
from pylibCZIrw import czi as pyczi
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROI_COLORS = ['#E63946', '#2196F3', '#FF9800', '#8BC34A',
              '#9C27B0', '#009688', '#FF5722', '#607D8B']

SUMMARY_COLORS = ['#E63946', '#2196F3', '#FF9800', '#8BC34A', '#B47CC7', '#555555']


DEFAULT_FRAP_CONFIG = {
    "frap_pattern": "*FRAP*.czi",
    "pixel_size_um": None,
    "imaging_bleach": True,
    "no_control": False, 
    "init": {
        "F_0": None,
        "f_bl": None,
        "f_mob": None,
        "D": 200,
        "t_b": None,
    },
    "bounds": {
        "F_0": [0, None],
        "f_bl": [0, 1.0],
        "f_mob": [0, 1.2],
        "D": [100, 1000],
        "t_b": [None, None],
    },
    "d_search_decades": 3,
    "outlier_z": 3.5,
}


def read_pixel_size_um(root):
    for dist in root.iter():
        if dist.tag.split('}')[-1] == 'Distance':
            if dist.get('Id', '') == 'X':
                for child in dist:
                    if child.tag.split('}')[-1] == 'Value' and child.text:
                        try:
                            val_m = float(child.text.strip())
                            if val_m > 0:
                                return val_m * 1e6
                        except ValueError:
                            pass
    return None

def normalise_without_control(raw_traces, bleach_frame):
    """
    Normalise each trace to its own pre-bleach mean.
    Returns norm_traces and ctrl_norm=None to signal no control was used.
    """
    norm_traces = []
    for tr in raw_traces:
        tr = tr.astype(float)
        pre_mean = float(np.nanmean(tr[:bleach_frame])) if bleach_frame > 0 else float(tr[0])
        if pre_mean <= 0:
            pre_mean = 1.0
        norm_traces.append(tr / pre_mean)
    return norm_traces, None


def normalise_with_control(raw_traces, ctrl_idx, bleach_frame):
    ctrl_trace = raw_traces[ctrl_idx].astype(float)
    ctrl_pre_mean = float(np.nanmean(ctrl_trace[:bleach_frame])) if bleach_frame > 0 else float(ctrl_trace[0])
    if ctrl_pre_mean <= 0:
        ctrl_pre_mean = 1.0

    ctrl_norm = ctrl_trace / ctrl_pre_mean
    ctrl_norm = np.clip(ctrl_norm, 0.01, np.inf)

    norm_traces = []
    for i, tr in enumerate(raw_traces):
        if i == ctrl_idx:
            norm_traces.append(tr.copy())
        else:
            norm_traces.append(tr.astype(float) / ctrl_norm)
    return norm_traces, ctrl_norm


def _soumpasis(x, x_0, R, D):
    dt = np.asarray(x, dtype=float) - x_0
    u = np.clip(R**2 / (2.0 * D * np.maximum(dt, 1e-12)), 0.0, 700.0)
    return np.exp(-2.0 * u) * (i0(u) + i1(u))


def model_with_ib(x, x_0, R, F_0, f_bl, f_mob, D, t_b):
    x = np.asarray(x, dtype=float)
    pre = x <= x_0
    y = np.empty_like(x)
    y[pre] = F_0 * np.exp(-x[pre] / t_b)
    S = _soumpasis(x[~pre], x_0, R, D)
    y[~pre] = np.exp(-x[~pre] / t_b) * (F_0 * (1 - f_bl) + F_0 * f_bl * f_mob * S)
    return y


def model_no_ib(x, x_0, R, F_0, f_bl, f_mob, D):
    x = np.asarray(x, dtype=float)
    pre = x <= x_0
    y = np.empty_like(x)
    y[pre] = F_0
    S = _soumpasis(x[~pre], x_0, R, D)
    y[~pre] = F_0 * (1 - f_bl) + F_0 * f_bl * f_mob * S
    return y


def evaluate_model(x, popt, imaging_bleach=True):
    if imaging_bleach:
        return model_with_ib(x, *popt)
    else:
        return model_no_ib(x, *popt)


def _estimate_D0(yf, bleach_frame, R_px):
    post = yf[bleach_frame:]
    if len(post) < 3:
        return 1.0
    lo = float(np.nanmin(post[:5]))
    hi = float(np.nanmax(post))
    target = lo + 0.5 * (hi - lo)
    cross = np.where(post >= target)[0]
    tau = (cross[0] / np.log(2)) if (cross.size > 0 and cross[0] > 0) else 1.0
    return float(R_px**2 / max(tau, 1e-3))


def fit_roi(norm_trace, bleach_frame, R_px, dt, pixel_size_um, config):
    imaging_bleach = config["imaging_bleach"]
    INIT = config["init"]
    BOUNDS = config["bounds"]
    D_SEARCH_DECADES = config["d_search_decades"]

    n = len(norm_trace)
    xf = np.arange(n, dtype=float)
    yf = norm_trace.astype(float)

    pre_mean = float(np.nanmean(yf[:bleach_frame])) if bleach_frame > 0 else float(yf[0])
    post_min = float(np.nanmin(yf[bleach_frame: bleach_frame + 5]))
    post_last = float(np.nanmean(yf[-max(1, n // 10):]))

    g_F0 = INIT['F_0'] if INIT['F_0'] is not None else max(pre_mean, 1.0)
    g_fbl = INIT['f_bl'] if INIT['f_bl'] is not None else float(np.clip((pre_mean - post_min) / (pre_mean + 1e-9), 0.05, 0.95))
    g_fmob = INIT['f_mob'] if INIT['f_mob'] is not None else float(np.clip((post_last - post_min) / (pre_mean - post_min + 1e-9), 0.1, 1.0))
    g_D = INIT['D'] if INIT['D'] is not None else _estimate_D0(yf, bleach_frame, R_px)
    g_tb = INIT['t_b'] if INIT['t_b'] is not None else float(n * 10)

    tb_lb = float(n * 2)

    def _b(key, fallback_lb, fallback_ub):
        lo = BOUNDS[key][0] if BOUNDS[key][0] is not None else fallback_lb
        hi = BOUNDS[key][1] if BOUNDS[key][1] is not None else fallback_ub
        return lo, hi

    F0_lb, F0_ub = _b('F_0', 0, np.inf)
    fbl_lb, fbl_ub = _b('f_bl', 0, 1.0)
    fmb_lb, fmb_ub = _b('f_mob', 0, 1.2)
    D_lb, D_ub = _b('D', 0, np.inf)

    if imaging_bleach:
        lb = [bleach_frame - 0.5, 0, F0_lb, fbl_lb, fmb_lb, D_lb, tb_lb]
        ub = [bleach_frame + 0.5, np.inf, F0_ub, fbl_ub, fmb_ub, D_ub, np.inf]
        model_fn = model_with_ib

        def make_p0(D_start):
            return [float(bleach_frame), R_px, g_F0, g_fbl, g_fmob, D_start, g_tb]
    else:
        lb = [bleach_frame - 0.5, 0, F0_lb, fbl_lb, fmb_lb, D_lb]
        ub = [bleach_frame + 0.5, np.inf, F0_ub, fbl_ub, fmb_ub, D_ub]
        model_fn = model_no_ib

        def make_p0(D_start):
            return [float(bleach_frame), R_px, g_F0, g_fbl, g_fmob, D_start]

    D_grid = np.logspace(
        np.log10(max(g_D, 1e-4)) - D_SEARCH_DECADES,
        np.log10(max(g_D, 1e-4)) + D_SEARCH_DECADES,
        2 * D_SEARCH_DECADES + 1
    ).clip(1e-4, 1e6)

    best_popt, best_pcov, best_ssr = None, None, np.inf
    for D_start in D_grid:
        try:
            po, pc = curve_fit(
                model_fn, xf, yf,
                p0=make_p0(float(D_start)),
                bounds=(lb, ub),
                maxfev=50000, ftol=1e-10, xtol=1e-10
            )
            ssr = float(np.sum((yf - model_fn(xf, *po))**2))
            if ssr < best_ssr:
                best_ssr, best_popt, best_pcov = ssr, po, pc
        except Exception:
            pass

    if best_popt is None:
        raise RuntimeError("All D starting points failed to converge.")

    R_fit, D_fit = best_popt[1], best_popt[5]
    tau_frames = R_fit**2 / D_fit
    tau_s = tau_frames * dt
    t_half_s = tau_s * np.log(2)
    D_um2s = D_fit * (pixel_size_um**2) / dt if pixel_size_um else None

    warn = ''
    if D_fit <= D_lb * (1 + 1e-3):
        warn = f'D at lower bound ({D_fit:.3g} px²/fr) — likely under-constrained'
    elif t_half_s > 5 * n * dt:
        warn = f't½={t_half_s:.2f} s >> acquisition length — D may be unreliable'

    return best_popt, best_pcov, t_half_s, tau_s, D_um2s, warn


def save_raw_excel(path, t_all, dt, rois, raw_traces, norm_traces,
                   ctrl_idx, frap_idxs, fit_results, bleach_frame, imaging_bleach):
    wb = Workbook()
    bold = Font(bold=True)
    n = len(t_all)

    ws = wb.active
    ws.title = 'Raw Intensity'
    hdrs = (['Frame', 'Time (s)']
            + [f"ROI {i+1} [Id={rois[i]['id']}]" + (' CTRL' if i == ctrl_idx else '')
               for i in range(len(rois))])
    for c, h in enumerate(hdrs, 1):
        ws.cell(1, c, h).font = bold
    for f, t in enumerate(t_all):
        r = f + 2
        ws.cell(r, 1, f)
        ws.cell(r, 2, round(float(t), 5))
        for c, tr in enumerate(raw_traces, 3):
            ws.cell(r, c, round(float(tr[f]), 4))

    ws_n = wb.create_sheet('Ctrl-Normalised + Fit')
    hdrs_n = (['Frame', 'Time (s)']
              + [f"ROI {i+1} [Id={rois[i]['id']}]" + (' CTRL' if i == ctrl_idx else '')
                 for i in range(len(rois))]
              + [f"ROI {frap_idxs[k]+1} fit (model)"
                 for k in range(len(frap_idxs)) if fit_results[k] is not None])
    for c, h in enumerate(hdrs_n, 1):
        ws_n.cell(1, c, h).font = bold

    xf = np.arange(n, dtype=float)
    fit_curves = {}
    for k, fi in enumerate(frap_idxs):
        if fit_results[k] is not None:
            fit_curves[k] = evaluate_model(xf, fit_results[k][0], imaging_bleach)

    n_rois = len(rois)
    fit_col_start = 3 + n_rois
    for f, t in enumerate(t_all):
        r = f + 2
        ws_n.cell(r, 1, f)
        ws_n.cell(r, 2, round(float(t), 5))
        for c, tr in enumerate(norm_traces, 3):
            ws_n.cell(r, c, round(float(tr[f]), 4))
        fit_col = fit_col_start
        for k in range(len(frap_idxs)):
            if k in fit_curves:
                ws_n.cell(r, fit_col, round(float(fit_curves[k][f]), 4))
                fit_col += 1

    for sheet in [ws, ws_n]:
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max(
                max(len(str(c.value or '')) for c in col), 8
            ) + 2

    wb.save(path)


def save_summary_excel(path, rois, frap_idxs, fit_results,
                       bleach_frame, dt, czi_name, pixel_size_um, imaging_bleach, ctrl_idx=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'FRAP Summary'
    bold = Font(bold=True)

    ws['A1'] = f'FRAP Summary — {czi_name}'
    ws['A1'].font = Font(bold=True, size=12)

    px_label = f'{pixel_size_um:.6f} µm' if pixel_size_um else 'not found in metadata'
    ctrl_label = (
        'NONE — each trace normalised to own pre-bleach mean'
        if ctrl_idx is None
        else 'YES — each FRAP trace divided by ctrl/ctrl_pre_mean'
    )

    meta = [
        ('Bleach frame', bleach_frame),
        ('Bleach time (s)', round(bleach_frame * dt, 4)),
        ('Frame interval (s)', round(dt, 5)),
        ('Model', 'Diff_2D_InfReservoir' + (' + IB' if imaging_bleach else ', no IB')),
        ('Control normalisation', ctrl_label),
        ('Pixel size (µm)', px_label),
    ]
    for i, (lbl, val) in enumerate(meta, 2):
        ws.cell(i, 1, lbl).font = bold
        ws.cell(i, 2, val)

    HR = len(meta) + 3
    hdrs = ['ROI #', 'Circle Id', 'cx (px)', 'cy (px)', 'R (px)',
            'F_0', 'f_bl', 'f_mob', 'Immobile',
            'D (px²/fr)', 'D (µm²/s)', 'τ (s)', 't½ (s)', 't_b (fr)', 'Warnings']
    for c, h in enumerate(hdrs, 1):
        ws.cell(HR, c, h).font = bold

    for i, (fi, res) in enumerate(zip(frap_idxs, fit_results)):
        row = HR + 1 + i
        roi = rois[fi]
        ws.cell(row, 1, f'ROI {fi+1}')
        ws.cell(row, 2, roi['id'])
        ws.cell(row, 3, round(roi['cx'], 2))
        ws.cell(row, 4, round(roi['cy'], 2))
        ws.cell(row, 5, round(roi['r'], 2))
        if res is not None:
            popt, pcov, t_half, tau_s, D_um2s, warn = res
            x0, R, F0, fbl, fmob, D = popt[0], popt[1], popt[2], popt[3], popt[4], popt[5]
            tb = popt[6] if imaging_bleach else float('nan')
            ws.cell(row, 6, round(float(F0), 2))
            ws.cell(row, 7, round(float(fbl), 4))
            ws.cell(row, 8, round(float(fmob), 4))
            ws.cell(row, 9, round(1 - float(fmob), 4))
            ws.cell(row, 10, round(float(D), 6))
            ws.cell(row, 11, round(float(D_um2s), 6) if D_um2s else 'n/a')
            ws.cell(row, 12, round(float(tau_s), 4))
            ws.cell(row, 13, round(float(t_half), 4))
            ws.cell(row, 14, round(float(tb), 2) if not np.isnan(tb) else 'n/a')
            ws.cell(row, 15, warn or 'OK')
        else:
            for c in range(6, len(hdrs) + 1):
                ws.cell(row, c, 'FAILED')

    nd = len(frap_idxs)
    LR = HR + nd
    SR = LR + 2
    for lbl, r in [('Mean', SR), ('SD', SR + 1), ('N', SR + 2)]:
        ws.cell(r, 1, lbl).font = bold
    col_letters = {7: 'G', 8: 'H', 9: 'I', 10: 'J', 12: 'L', 13: 'M'}
    for ci, cl in col_letters.items():
        rng = f'{cl}{HR+1}:{cl}{LR}'
        ws.cell(SR, ci, f'=IFERROR(AVERAGE({rng}),"")').number_format = '0.0000'
        ws.cell(SR + 1, ci, f'=IFERROR(STDEV({rng}),"")').number_format = '0.0000'
    ws.cell(SR + 2, 7, nd)

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            max(len(str(c.value or '')) for c in col), 8
        ) + 2

    wb.save(path)


def mad_filter(values, z_thresh=3.5):
    values = np.asarray(values, dtype=float)
    if len(values) < 3:
        return np.ones(len(values), dtype=bool)
    median = np.median(values)
    mad = np.median(np.abs(values - median))
    if mad == 0:
        return np.ones(len(values), dtype=bool)
    robust_z = 0.6745 * (values - median) / mad
    return np.abs(robust_z) <= z_thresh


def combined_outlier_mask(t, m, d, z_thresh=3.5):
    return mad_filter(t, z_thresh) & mad_filter(m, z_thresh) & mad_filter(d, z_thresh)


def read_summary_xlsx(xlsx_path: Path):
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    header_row = None
    col_map = {}
    for r in range(1, 30):
        row_vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
        if "ROI #" in row_vals:
            header_row = r
            for i, name in enumerate(row_vals):
                col_map[name] = i + 1
            break

    if header_row is None:
        raise ValueError("Header row not found")

    required = ["ROI #", "f_mob", "t½ (s)", "D (µm²/s)"]
    for k in required:
        if k not in col_map:
            raise ValueError(f"Missing column: {k}")

    t_halfs, mobiles, diffs = [], [], []
    row = header_row + 1
    while True:
        roi = ws.cell(row=row, column=col_map["ROI #"]).value
        if roi is None or str(roi).strip() in ("Mean", "SD", "N", ""):
            break
        try:
            mobiles.append(float(ws.cell(row=row, column=col_map["f_mob"]).value))
            t_halfs.append(float(ws.cell(row=row, column=col_map["t½ (s)"]).value))
            diffs.append(float(ws.cell(row=row, column=col_map["D (µm²/s)"]).value))
        except (TypeError, ValueError):
            pass
        row += 1

    if not t_halfs:
        raise ValueError("No valid ROI rows")

    return np.mean(t_halfs), np.mean(mobiles), np.mean(diffs)


def collect_condition_data(head: Path, outlier_z=3.5):
    results = {}
    for cond_dir in sorted(d for d in head.iterdir() if d.is_dir()):
        t_all, m_all, d_all = [], [], []
        for rep_dir in sorted(d for d in cond_dir.iterdir() if d.is_dir()):
            summaries = list(rep_dir.rglob('*_FRAP_summary.xlsx'))
            if not summaries:
                continue
            try:
                t, m, d = read_summary_xlsx(summaries[0])
                t_all.append(t)
                m_all.append(m)
                d_all.append(d)
            except Exception:
                pass

        if len(t_all) < 3:
            continue

        mask = combined_outlier_mask(t_all, m_all, d_all, z_thresh=outlier_z)
        results[cond_dir.name] = {
            't_half': np.array(t_all)[mask].tolist(),
            'mobile': np.array(m_all)[mask].tolist(),
            'diff': np.array(d_all)[mask].tolist(),
        }
    return results


def analyse_frap(czi_path, config=None):
    config = DEFAULT_FRAP_CONFIG.copy() if config is None else config
    czi_path = Path(czi_path)

    with pyczi.open_czi(str(czi_path)) as czidoc:
        bbox = czidoc.total_bounding_box
        n_frames = bbox.get('T', (0, 1))[1]
        frames = []
        for t in range(n_frames):
            plane = czidoc.read(plane={'T': t, 'C': 0, 'Z': 0})
            frames.append(np.squeeze(plane))
        data = np.stack(frames, axis=0)
        root = ET.fromstring(czidoc.raw_metadata)

    dt = None
    for el in root.iter():
        if el.tag.split('}')[-1] == 'FrameTime' and el.text:
            try:
                v = float(el.text.strip())
                if v > 0:
                    dt = v
                    break
            except Exception:
                pass
    dt = dt or 1.0

    pixel_size_um = read_pixel_size_um(root)
    if pixel_size_um is None:
        pixel_size_um = config.get("pixel_size_um", None)

    # ── Get ROIs ──
    rois = []
    for circle in root.iter():
        if circle.tag.split('}')[-1] != 'Circle':
            continue
        geom = next((c for c in circle if c.tag.split('}')[-1] == 'Geometry'), None)
        if geom is None:
            continue

        def _v(p, tag):
            for c in p:
                if c.tag.split('}')[-1] == tag and c.text:
                    try:
                        return float(c.text.strip())
                    except Exception:
                        pass
            return None

        cx, cy, r = _v(geom, 'CenterX'), _v(geom, 'CenterY'), _v(geom, 'Radius')
        if None not in (cx, cy, r):
            rois.append({'id': circle.get('Id'), 'cx': cx, 'cy': cy, 'r': r})

    # ── Validate ROI count ──
    n_rois_expected = config.get("n_rois", None)
    no_control = config.get("no_control", False)

    if n_rois_expected is not None:
        if len(rois) < n_rois_expected:
            raise RuntimeError(
                f"Expected {n_rois_expected} ROIs but found only {len(rois)} in metadata. "
                f"Check that the CZI file has the correct ROI annotations."
            )
        if len(rois) > n_rois_expected:
            # use only the first n_rois_expected ROIs
            rois = rois[:n_rois_expected]

    if not no_control and len(rois) < 2:
        raise RuntimeError(f"{len(rois)} ROI(s) found — need at least 2 (1 FRAP + 1 control).")
    if no_control and len(rois) < 1:
        raise RuntimeError("No ROIs found in metadata.")

    # ── Extract traces ──
    def trace(cx, cy, r):
        rr, cc = disk((cy, cx), r, shape=data.shape[1:])
        return data[:, rr, cc].mean(axis=1).astype(float)

    raw_traces = [trace(roi['cx'], roi['cy'], roi['r']) for roi in rois]

    # ── Detect bleach frame ──
    drop = np.zeros(n_frames - 1)
    for tr in raw_traces:
        drop += np.maximum(tr[:-1] - tr[1:], 0)
    bleach_frame = int(np.argmax(drop)) + 1

    # ── Identify control ROI ──
    user_ctrl_idx = config.get("ctrl_idx", None)

    if no_control:
        ctrl_idx = None
        frap_idxs = list(range(len(raw_traces)))
    elif user_ctrl_idx is not None:
        if user_ctrl_idx < 0 or user_ctrl_idx >= len(rois):
            raise RuntimeError(
                f"Control ROI index {user_ctrl_idx} is out of range. "
                f"Valid range: 0 to {len(rois) - 1}."
            )
        ctrl_idx = user_ctrl_idx
        frap_idxs = [i for i in range(len(raw_traces)) if i != ctrl_idx]
    else:
        # auto-detect: ROI with smallest bleach drop
        def fdrop(tr, bf):
            return (
                (np.mean(tr[:bf]) - np.mean(tr[bf:bf + 3]))
                / (np.mean(tr[:bf]) + 1e-9)
            )

        drops = [fdrop(tr, bleach_frame) for tr in raw_traces]
        ctrl_idx = int(np.argmin(drops))
        frap_idxs = [i for i in range(len(raw_traces)) if i != ctrl_idx]

    # ── Normalise ──
    if no_control or ctrl_idx is None:
        norm_traces, ctrl_norm = normalise_without_control(raw_traces, bleach_frame)
    else:
        norm_traces, ctrl_norm = normalise_with_control(
            raw_traces, ctrl_idx, bleach_frame
        )

    
    colors = ROI_COLORS[:len(frap_idxs)]
    t_all = np.arange(n_frames) * dt

    fit_results = []
    for fi in frap_idxs:
        try:
            res = fit_roi(norm_traces[fi], bleach_frame, rois[fi]['r'], dt, pixel_size_um, config)
            fit_results.append(res)
        except Exception:
            fit_results.append(None)

    # Save xlsx outputs only; figure is built in GUI only
    xls1 = czi_path.with_name(czi_path.stem + '_FRAP_raw_data.xlsx')
    save_raw_excel(
        xls1, t_all, dt, rois, raw_traces, norm_traces,
        ctrl_idx, frap_idxs, fit_results, bleach_frame, config["imaging_bleach"]
    )

    xls2 = czi_path.with_name(czi_path.stem + '_FRAP_summary.xlsx')
    save_summary_excel(
        xls2, rois, frap_idxs, fit_results,
        bleach_frame, dt, czi_path.name, pixel_size_um, config["imaging_bleach"],ctrl_idx=ctrl_idx,
    )

    return {
        "czi_path": str(czi_path),
        "stem": czi_path.stem,
        "raw_xlsx": str(xls1),
        "summary_xlsx": str(xls2),
        "bleach_frame": bleach_frame,
        "dt": dt,
        "pixel_size_um": pixel_size_um,
        "ctrl_idx": ctrl_idx,
        "no_control": no_control,
        "ctrl_idx_source": (
            "none"  if no_control
            else "user" if user_ctrl_idx is not None
            else "auto"
        ),
        "frap_idxs": frap_idxs,
        "rois": rois,
        "n_rois_in_metadata": len(rois),
        "fit_results": fit_results,
        "t_all": t_all.tolist(),
        "raw_traces": [tr.tolist() for tr in raw_traces],
        "norm_traces": [tr.tolist() for tr in norm_traces],
        "roi_colors": colors,
        "imaging_bleach": config["imaging_bleach"],
    }


def run_frap_batch(folder, pattern="*FRAP*.czi", config=None, progress_queue=None, cancel_event=None):
    root = Path(folder)
    hits = sorted(root.rglob(pattern))
    results = []
    failed = []

    total = len(hits)
    if progress_queue is not None:
        progress_queue.put(("progress", 0.0))

    for i, p in enumerate(hits):
        if cancel_event is not None and cancel_event.is_set():
            break
        try:
            res = analyse_frap(p, config=config)
            results.append(res)
            if progress_queue is not None:
                progress_queue.put(("file_done", res))
                progress_queue.put(("progress", 100.0 * (i + 1) / max(1, total)))
        except Exception:
            failed.append(str(p))
            if progress_queue is not None:
                progress_queue.put(("progress", 100.0 * (i + 1) / max(1, total)))

    return {
        "n_total": total,
        "n_ok": len(results),
        "n_failed": len(failed),
        "failed": failed,
        "last_res": results[-1] if results else None,
    }